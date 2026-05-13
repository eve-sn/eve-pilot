"""
Import du portefeuille projets depuis Budget_Previsionnel_EVE_2026.xlsx.

Source: onglet 6 "Portefeuille projets". Seul le portefeuille est traite ici ;
les onglets charges fixes / IPRES-CSS / VRS-BRS / apurement / plan tresorerie
sont laisses pour des commandes dediees ulterieures.

Decisions metier figees a date du 13/05/2026:
- AGIR Pikine Phase I exclu (cloture fevrier 2026, deja passe).
- Pikine Phase II -> projet existant NOUSCIMS-PIK-MBAO-2026 (update).
- Saint-Louis Gouvernance -> projet existant NOUSCIMS-SL-2026 renomme (le projet
  base RH s'appelait "Nutrition Saint-Louis", arbitrage utilisateur: meme projet
  recadre cote document budget).
- Les autres projets (Espaces communautaires, PDBH IEC, Inondations, ISF, ECO-AVENIR)
  sont crees ex-nihilo avec un code derive.
- PNBSF-DAK-2026 et PNBSF-KED-2026 passent en status PREPARATION (offres de
  service non encore signees).

La commande est idempotente: relance sans duplication via update_or_create.
"""

from datetime import date
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.projects.models import Donor, Project


XLSX_REL_PATH = "Budget_Previsionnel_EVE_2026.xlsx"
XLSX_SHEET = "6. Portefeuille projets"
SOURCE_LABEL = "Budget Previsionnel EVE 2026 - onglet 6 Portefeuille projets"


NEW_DONORS = {
    "ONAS-AFD": {
        "short_name": "ONAS-AFD",
        "donor_type": Donor.DonorType.BILATERAL,
        "country": "Senegal / France",
    },
    "ChildFund / P&G": {
        "short_name": "ChildFund",
        "donor_type": Donor.DonorType.FOUNDATION,
        "country": "International",
    },
    "AXA Climate": {
        "short_name": "AXA Climate",
        "donor_type": Donor.DonorType.COMPANY,
        "country": "France",
    },
    "Oghogho Meye / La Locomotiva": {
        "short_name": "La Locomotiva",
        "donor_type": Donor.DonorType.FOUNDATION,
        "country": "Italie",
    },
}


# Specs projets a importer. Une entree par ligne du xlsx onglet 6.
# Le 'xlsx_label' doit matcher exactement la colonne Projet du xlsx pour relier
# les valeurs Budget total / % Exec lues a l'execution.
PROJECT_SPECS = [
    {
        "xlsx_label": "Pikine Phase II — Nutrition",
        "code": "NOUSCIMS-PIK-MBAO-2026",
        "title": "Pikine Phase II - Nutrition (Nous-Cims)",
        "short_title": "Pikine Phase II Nutrition",
        "sector": "NUTRITION",
        "start_date": date(2026, 1, 1),
        "end_date": date(2027, 12, 31),
        "donor_name": "Fondation Nous-Cims",
        "description": (
            "Source: " + SOURCE_LABEL + ". Couvre les tranches T1 (70k EUR) "
            "et T2 (60k EUR) selon le plan de tresorerie 2026."
        ),
    },
    {
        "xlsx_label": "Saint-Louis — Gouvernance Multisectorielle",
        "code": "NOUSCIMS-SL-2026",
        "title": "Saint-Louis - Gouvernance Multisectorielle (Nous-Cims)",
        "short_title": "Saint-Louis Gouvernance",
        "sector": "GOUVERNANCE",
        "start_date": date(2025, 8, 1),
        "end_date": date(2028, 7, 31),
        "donor_name": "Fondation Nous-Cims",
        "description": (
            "Source: " + SOURCE_LABEL + ". Projet renomme d'apres le document "
            "budget (precedemment intitule Nutrition Saint-Louis dans la base RH)."
        ),
    },
    {
        "xlsx_label": "Espaces communautaires Pikine",
        "code": "NOUSCIMS-ECP-2025",
        "title": "Espaces communautaires Pikine (Nous-Cims)",
        "short_title": "Espaces communautaires Pikine",
        "sector": "COMMUNAUTAIRE",
        "start_date": date(2025, 6, 1),
        "end_date": date(2026, 5, 31),
        "donor_name": "Fondation Nous-Cims",
        "description": "Source: " + SOURCE_LABEL + ". Cloture prevue mai 2026.",
    },
    {
        "xlsx_label": "PDBH IEC — Avenant 3 + soldes",
        "code": "ONASAFD-PDBH-IEC-2025",
        "title": "PDBH IEC - Avenant 3 + soldes (ONAS-AFD)",
        "short_title": "PDBH IEC",
        "sector": "EAU_ASSAINISSEMENT",
        "start_date": date(2024, 9, 1),
        "end_date": date(2028, 5, 31),
        "donor_name": "ONAS-AFD",
        "description": (
            "Source: " + SOURCE_LABEL + ". Periode '44 mois' interpretee comme "
            "Sep-2024 -> Mai-2028, a confirmer cote contrat."
        ),
    },
    {
        "xlsx_label": "Réponse urgence inondations",
        "code": "CHILDFUND-INONDATIONS-2025",
        "title": "Reponse urgence inondations (ChildFund / P&G)",
        "short_title": "Inondations",
        "sector": "URGENCE",
        "start_date": date(2025, 12, 1),
        "end_date": date(2026, 3, 31),
        "donor_name": "ChildFund / P&G",
        "description": (
            "Source: " + SOURCE_LABEL + ". Periode xlsx Dec-2025 -> Mars-2026. "
            "Projet declare encore actif a la date d'import malgre date de fin "
            "depassee : prorogation ou avenant a confirmer cote bailleur."
        ),
    },
    {
        "xlsx_label": "ISF / AXA Climate",
        "code": "AXA-ISF-2026",
        "title": "ISF / AXA Climate (94 JH + frais - reliquat 2026)",
        "short_title": "ISF AXA Climate",
        "sector": "CLIMAT",
        "start_date": date(2026, 1, 1),
        "end_date": date(2026, 7, 31),
        "donor_name": "AXA Climate",
        "description": (
            "Source: " + SOURCE_LABEL + ". Reliquat 2026 d'un projet 2025 "
            "(7,59M FCFA deja decaisses). Periode estimee depuis le plan de "
            "tresorerie mars-juillet 2026."
        ),
    },
    {
        "xlsx_label": "ECO-AVENIR (11 500 €)",
        "code": "OGHOGHO-ECOAVENIR-2026",
        "title": "ECO-AVENIR (Oghogho Meye / La Locomotiva)",
        "short_title": "ECO-AVENIR",
        "sector": "ENVIRONNEMENT",
        "start_date": date(2026, 1, 1),
        "end_date": date(2026, 12, 31),
        "donor_name": "Oghogho Meye / La Locomotiva",
        "description": (
            "Source: " + SOURCE_LABEL + ". Budget 11 500 EUR sur 12 mois "
            "(equivalent xlsx 7 543 506 FCFA)."
        ),
    },
]


# Projets sans contrat signe a date: bascule en PREPARATION sans toucher au RH.
PROJECTS_TO_PREPARATION = ["PNBSF-DAK-2026", "PNBSF-KED-2026"]


class Command(BaseCommand):
    help = "Importe le portefeuille projets depuis Budget_Previsionnel_EVE_2026.xlsx (onglet 6)."

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(
                "openpyxl manquant. Installer via 'pip install openpyxl' avant relance."
            )
            return

        xlsx_path = Path(settings.BASE_DIR) / XLSX_REL_PATH
        if not xlsx_path.exists():
            self.stderr.write(f"Fichier xlsx introuvable: {xlsx_path}")
            return

        self.stdout.write(f"Lecture {xlsx_path.name} (onglet '{XLSX_SHEET}')...")
        xlsx_rows = self._load_portfolio_rows(openpyxl, xlsx_path)

        donors = self._seed_donors()
        self._import_projects(xlsx_rows, donors)
        self._mark_preparation()

        self.stdout.write(self.style.SUCCESS("Import budget portefeuille termine."))

    def _load_portfolio_rows(self, openpyxl, xlsx_path):
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        ws = wb[XLSX_SHEET]
        rows_by_label = {}
        # Donnees a partir de la ligne 4 (entete en ligne 3). On STOP des que la colonne
        # numero n'est plus un entier (= on quitte la zone projets pour la ligne TOTAL
        # ou la section "Encaissements documentes" qui suit). Sans ce garde-fou, certains
        # libelles d'encaissement reutilisent les memes noms de projet et ecrasent
        # silencieusement les budgets.
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            num, project, donor, period, total_budget, realized, remainder, exec_rate = row[:8]
            if not isinstance(num, int):
                break
            if not project:
                continue
            label = str(project).strip()
            rows_by_label[label] = {
                "donor_label": str(donor).strip() if donor else "",
                "period_label": str(period).strip() if period else "",
                "total_budget": self._to_decimal(total_budget, label, "total_budget"),
                "realized": self._to_decimal(realized, label, "realized"),
                "remainder_2026": self._to_decimal(remainder, label, "remainder"),
                "exec_rate": self._to_decimal(exec_rate, label, "exec_rate"),
            }
        self.stdout.write(f"  {len(rows_by_label)} lignes portefeuille extraites du xlsx.")
        return rows_by_label

    def _to_decimal(self, value, label, field):
        """Convertit une cellule xlsx en Decimal ou None, en signalant les valeurs non numeriques."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip()
        if not text:
            return None
        try:
            return Decimal(text.replace(" ", "").replace(",", "."))
        except Exception:
            self.stderr.write(
                f"  /!\\ Valeur non numerique pour '{label}'.{field} = {value!r}, ignoree."
            )
            return None

    def _seed_donors(self):
        existing = {d.name: d for d in Donor.objects.filter(is_active=True, deleted_at__isnull=True)}
        for name, defaults in NEW_DONORS.items():
            donor, created = Donor.objects.update_or_create(name=name, defaults=defaults)
            action = "cree" if created else "deja present"
            self.stdout.write(f"  Donor '{name}': {action}")
            existing[name] = donor
        return existing

    def _import_projects(self, xlsx_rows, donors):
        for spec in PROJECT_SPECS:
            xlsx = xlsx_rows.get(spec["xlsx_label"])
            if xlsx is None:
                self.stderr.write(
                    f"  /!\\ Ligne xlsx introuvable pour '{spec['xlsx_label']}', projet ignore."
                )
                continue
            donor = donors.get(spec["donor_name"])
            if donor is None:
                self.stderr.write(
                    f"  /!\\ Donor '{spec['donor_name']}' introuvable pour projet {spec['code']}."
                )
                continue

            exec_rate = xlsx["exec_rate"] or Decimal("0")
            progress = (exec_rate * Decimal("100")).quantize(Decimal("0.01"))
            if progress > Decimal("100.00"):
                progress = Decimal("100.00")

            defaults = {
                "title": spec["title"],
                "short_title": spec["short_title"],
                "primary_donor": donor,
                "total_budget": xlsx["total_budget"],
                "currency": "XOF",
                "start_date": spec["start_date"],
                "end_date": spec["end_date"],
                "status": Project.Status.ACTIVE,
                "sector": spec["sector"],
                "progress_percentage": progress,
                "description": spec["description"],
            }
            project, created = Project.objects.update_or_create(
                code=spec["code"], defaults=defaults
            )
            action = "cree" if created else "mis a jour"
            self.stdout.write(
                f"  Project {spec['code']}: {action} "
                f"(budget={defaults['total_budget']}, progress={progress}%)"
            )

    def _mark_preparation(self):
        for code in PROJECTS_TO_PREPARATION:
            try:
                project = Project.objects.get(
                    code=code, is_active=True, deleted_at__isnull=True
                )
            except Project.DoesNotExist:
                self.stderr.write(f"  /!\\ Project {code} introuvable, bascule ignoree.")
                continue
            if project.status == Project.Status.PREPARATION:
                self.stdout.write(f"  Project {code}: deja en PREPARATION")
                continue
            previous = project.status
            project.status = Project.Status.PREPARATION
            project.save(update_fields=["status", "updated_at"])
            self.stdout.write(f"  Project {code}: {previous} -> PREPARATION")
